#!/usr/bin/env python3
"""Конвертирует Diadoc_status.xlsx → data/dashboard.json для GitHub Pages."""

import json, re, os
from openpyxl import load_workbook
from datetime import datetime

XLSX_PATH = os.path.join(os.path.dirname(__file__), "Diadoc_status.xlsx")
DZO_PATH = os.path.join(os.path.dirname(__file__), "dzo_inn.json")
OUT_PATH = os.path.join(os.path.dirname(__file__), "data", "dashboard.json")

STATUS_MAP = {
    "Требуется подписать и отправить. На подписании": "Требуется подписать и отправить",
}


def clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    m = re.match(r'^="(.+)"$', s)
    return m.group(1) if m else s


def doc_type(doc_num):
    s = str(doc_num) if doc_num else ""
    if "LBO" in s:
        return "PAYG"
    if "LBS" in s:
        return "Пакеты"
    return "Другое"


def normalize_status(status):
    # Сначала проверяем прямое соответствие
    if status in STATUS_MAP:
        return STATUS_MAP[status]
    # "Подписан всеми получателями. На согласовании" → "Подписан всеми получателями"
    if status.startswith("Подписан всеми получателями"):
        return "Подписан всеми получателями"
    return status


def load_dzo_inns():
    if os.path.exists(DZO_PATH):
        with open(DZO_PATH, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb.active
    dzo_inns = load_dzo_inns()

    statuses = {}
    types = {"PAYG": 0, "Пакеты": 0}
    segments = {"ДЗО": 0, "Внешние": 0}
    type_segment = {"PAYG|ДЗО": 0, "PAYG|Внешние": 0, "Пакеты|ДЗО": 0, "Пакеты|Внешние": 0}
    status_by_type = {}
    status_by_segment = {}
    signed = []
    total = 0

    for row in range(2, ws.max_row + 1):
        inn = clean(ws.cell(row=row, column=1).value)
        org = clean(ws.cell(row=row, column=3).value)
        doc_num = clean(ws.cell(row=row, column=5).value)
        raw_status = clean(ws.cell(row=row, column=14).value)
        date_changed = clean(ws.cell(row=row, column=16).value)

        if not org:
            continue

        total += 1
        status = normalize_status(raw_status)
        dtype = doc_type(doc_num)
        segment = "ДЗО" if inn in dzo_inns else "Внешние"

        statuses[status] = statuses.get(status, 0) + 1
        if dtype in types:
            types[dtype] = types.get(dtype, 0) + 1
        segments[segment] = segments.get(segment, 0) + 1

        ts_key = f"{dtype}|{segment}"
        if ts_key in type_segment:
            type_segment[ts_key] += 1

        key_t = f"{dtype}|{status}"
        status_by_type[key_t] = status_by_type.get(key_t, 0) + 1

        key_s = f"{segment}|{status}"
        status_by_segment[key_s] = status_by_segment.get(key_s, 0) + 1

        if status == "Подписан всеми получателями":
            # Извлекаем только дату ДД.ММ.ГГГГ из "20.04.2026 18:04:08"
            sign_date = date_changed.split(" ")[0] if date_changed else ""
            signed.append({
                "type": dtype,
                "segment": segment,
                "date": sign_date,
            })

    all_statuses = sorted(statuses.keys())
    cross = []
    for s in all_statuses:
        cross.append({
            "status": s,
            "PAYG": status_by_type.get(f"PAYG|{s}", 0),
            "Пакеты": status_by_type.get(f"Пакеты|{s}", 0),
            "ДЗО": status_by_segment.get(f"ДЗО|{s}", 0),
            "Внешние": status_by_segment.get(f"Внешние|{s}", 0),
        })

    result = {
        "updated": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "total": total,
        "statuses": statuses,
        "types": types,
        "segments": segments,
        "type_segment": {
            "PAYG_ДЗО": type_segment["PAYG|ДЗО"],
            "PAYG_Внешние": type_segment["PAYG|Внешние"],
            "Пакеты_ДЗО": type_segment["Пакеты|ДЗО"],
            "Пакеты_Внешние": type_segment["Пакеты|Внешние"],
        },
        "cross": cross,
        "signed": sorted(signed, key=lambda x: x["date"]),
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"OK: {total} строк → {OUT_PATH}")


if __name__ == "__main__":
    main()
